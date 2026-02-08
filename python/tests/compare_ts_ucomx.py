"""Compare NEW TypeScript metrics (after fix) against UCoMx reference."""
import json
import openpyxl
import os

# Load TS metrics
ts_path = r'C:\Users\teoir\OneDrive\Desktop\rt-complexity-lens\python\tests\reference_data\reference_metrics_ts.json'
with open(ts_path) as f:
    ts_data = json.load(f)

# Load UCoMx reference
wb = openpyxl.load_workbook(
    r'C:\Users\teoir\OneDrive\Desktop\rt-complexity-lens\testdata\reference_dataset_v1.1\0-all-20262822356.397\dataset.xlsx'
)
ws = wb.active
headers = [c.value for c in ws[1]]
ucomx_rows = []
for row in ws.iter_rows(min_row=2):
    d = {h: row[i].value for i, h in enumerate(headers)}
    ucomx_rows.append(d)

# Metric mapping: TS name -> UCoMx column name
metric_map = {
    'totalMU': 'MUs',
    'LSV': 'LSV',
    'AAV': 'AAV',
    'MCS': 'MCSv',
}

# Match TS plans to UCoMx by total MU
ts_plans = ts_data.get('plans', ts_data) if isinstance(ts_data, dict) else ts_data
if isinstance(ts_plans, dict):
    ts_plans = list(ts_plans.values()) if not 'plans' in ts_plans else ts_plans['plans']

# Build filename -> MU mapping for TS
ts_by_mu = {}
for p in ts_plans:
    mu = p.get('totalMU', 0)
    ts_by_mu[round(mu, 2)] = p

print(f"Loaded {len(ts_plans)} TS plans, {len(ucomx_rows)} UCoMx plans")
print()

# Compare
print(f"{'Plan':<18} {'Metric':<8} {'TS':>10} {'UCoMx':>10} {'Ratio':>8} {'Delta%':>8}")
print("-" * 72)

matches = 0
mismatches = 0
close_matches = 0

for ref in ucomx_rows:
    ref_mu = ref['MUs']
    
    # Find TS plan with matching MU
    ts = None
    for mu_key, plan in ts_by_mu.items():
        if abs(mu_key - ref_mu) < 1:
            ts = plan
            break
    
    if ts is None:
        continue
    
    plan_name = ts.get('planLabel', ts.get('fileName', f'MU={ref_mu}'))
    # Shorten to last part
    if '/' in plan_name:
        plan_name = plan_name.split('/')[-1]
    plan_name = plan_name.replace('.dcm', '')[:18]
    
    for ts_key, ucomx_key in metric_map.items():
        ts_val = ts.get(ts_key)
        ucomx_val = ref.get(ucomx_key)
        
        if ts_val is None or ucomx_val is None or ucomx_val == 0:
            continue
        
        ratio = ts_val / ucomx_val
        delta_pct = (ts_val - ucomx_val) / ucomx_val * 100
        
        status = '✓' if abs(delta_pct) < 5 else ('~' if abs(delta_pct) < 15 else '✗')
        if abs(delta_pct) < 5:
            close_matches += 1
        elif abs(delta_pct) < 15:
            matches += 1
        else:
            mismatches += 1
        
        print(f"{plan_name:<18} {ts_key:<8} {ts_val:10.4f} {ucomx_val:10.4f} {ratio:8.4f} {delta_pct:+8.1f}% {status}")

total = close_matches + matches + mismatches
print(f"\n--- Summary ---")
print(f"Close (<5%):  {close_matches}")
print(f"Near (<15%):  {matches}")
print(f"Far (>15%):   {mismatches}")
print(f"Total:        {total}")
