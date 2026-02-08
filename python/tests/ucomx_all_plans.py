"""Compare UCoMx metrics across multiple plans to find AAV pattern."""
import pydicom
import numpy as np
import openpyxl
import os
import glob

# Load UCoMx reference
wb = openpyxl.load_workbook(
    r'C:\Users\teoir\OneDrive\Desktop\rt-complexity-lens\testdata\reference_dataset_v1.1\0-all-20262822356.397\dataset.xlsx'
)
ws = wb.active
headers = [c.value for c in ws[1]]
ucomx_rows = []
for row in ws.iter_rows(min_row=2):
    d = {h: row[i].value for i, h in enumerate(headers)}
    ucomx_rows.append(d)

# Find all plans
base = r'C:\Users\teoir\OneDrive\Desktop\rt-complexity-lens\testdata\reference_dataset_v1.1\Linac'
plans = sorted(glob.glob(os.path.join(base, '*', 'RTPLAN_*.dcm')))
print(f"Found {len(plans)} plans")


def lsv_bank(positions, active_mask):
    idx = np.where(active_mask)[0]
    if len(idx) < 2:
        return 1.0
    diffs = np.abs(np.diff(positions[idx]))
    mx = np.max(diffs)
    return np.mean(1.0 - diffs / mx) if mx > 0 else 1.0


def compute_metrics(dcm_path):
    """Compute LSV, AAV, MCS for a single plan."""
    ds = pydicom.dcmread(dcm_path)
    results = {}

    for beam in ds.BeamSequence:
        # Get leaf boundaries
        bounds = None
        n_pairs = None
        for bld in beam.BeamLimitingDeviceSequence:
            if bld.RTBeamLimitingDeviceType in ('MLCX', 'MLCY'):
                bounds = np.array(bld.LeafPositionBoundaries)
                n_pairs = bld.NumberOfLeafJawPairs
                mlc_type = bld.RTBeamLimitingDeviceType
                break
        if bounds is None:
            continue

        # Parse CPs
        cp_data = []
        prev_mlc = None
        prev_y = None
        prev_x = None
        all_gaps = []
        for cp in beam.ControlPointSequence:
            mlc = None
            y = prev_y
            x = prev_x
            if hasattr(cp, 'BeamLimitingDevicePositionSequence'):
                for bldp in cp.BeamLimitingDevicePositionSequence:
                    if bldp.RTBeamLimitingDeviceType in ('MLCX', 'MLCY'):
                        pos = np.array(bldp.LeafJawPositions, dtype=float)
                        half = len(pos) // 2
                        mlc = (pos[:half], pos[half:])
                    elif bldp.RTBeamLimitingDeviceType == 'ASYMY':
                        y = np.array(bldp.LeafJawPositions, dtype=float)
                    elif bldp.RTBeamLimitingDeviceType == 'ASYMX':
                        x = np.array(bldp.LeafJawPositions, dtype=float)
            if mlc is None:
                mlc = prev_mlc
            if y is None:
                y = prev_y
            if x is None:
                x = prev_x

            gaps = mlc[1] - mlc[0]
            all_gaps.extend(gaps.tolist())
            w = float(cp.CumulativeMetersetWeight) if hasattr(cp, 'CumulativeMetersetWeight') else 0
            cp_data.append({
                'a': mlc[0].copy(), 'b': mlc[1].copy(),
                'gaps': gaps.copy(),
                'y': y.copy() if y is not None else np.array([-200, 200]),
                'x': x.copy() if x is not None else None,
                'w': w
            })
            prev_mlc = mlc
            prev_y = y
            prev_x = x

        min_gap = min(all_gaps) if all_gaps else 0
        n_ca = len(cp_data) - 1
        if n_ca == 0:
            continue

        # Build CAs with midpoint interpolation
        areas = []
        lsv_cas = []
        for j in range(n_ca):
            cp1, cp2 = cp_data[j], cp_data[j + 1]
            ca_a = (cp1['a'] + cp2['a']) / 2
            ca_b = (cp1['b'] + cp2['b']) / 2
            ca_gaps = ca_b - ca_a
            ca_y = (cp1['y'] + cp2['y']) / 2

            # Active leaf mask
            active = np.zeros(n_pairs, dtype=bool)
            area = 0
            for k in range(n_pairs):
                within_jaw = (bounds[k + 1] > ca_y[0]) and (bounds[k] < ca_y[1])
                if within_jaw and ca_gaps[k] > min_gap:
                    active[k] = True
                    eff_w = max(0, min(bounds[k + 1], ca_y[1]) - max(bounds[k], ca_y[0]))
                    area += ca_gaps[k] * eff_w
            areas.append(area)

            la = lsv_bank(ca_a, active)
            lb = lsv_bank(ca_b, active)
            lsv_cas.append((la + lb) / 2)

        areas = np.array(areas)
        lsv_cas = np.array(lsv_cas)
        a_max = np.max(areas) if len(areas) > 0 else 1
        aav_cas = areas / a_max
        mcs_cas = lsv_cas * aav_cas

        beam_name = getattr(beam, 'BeamName', str(beam.BeamNumber))
        results[beam_name] = {
            'lsv': np.mean(lsv_cas),
            'aav': np.mean(aav_cas),
            'mcs': np.mean(mcs_cas),
            'n_ca': n_ca,
            'a_max': a_max,
            'mean_area': np.mean(areas),
            'min_gap': min_gap,
        }

    # Average across beams (equation 1)
    if results:
        return {
            'lsv': np.mean([r['lsv'] for r in results.values()]),
            'aav': np.mean([r['aav'] for r in results.values()]),
            'mcs': np.mean([r['mcs'] for r in results.values()]),
            'n_beams': len(results),
            'a_max': max(r['a_max'] for r in results.values()),
            'min_gap': min(r['min_gap'] for r in results.values()),
        }
    return None


# Run comparison
print(f"{'Plan':<16} {'LSV_c':>7} {'LSV_u':>7} {'ratio':>7} | {'AAV_c':>7} {'AAV_u':>7} {'ratio':>7} | {'MCS_c':>7} {'MCS_u':>7} {'ratio':>7}")
print("-" * 100)

for plan_path in plans:
    name = os.path.basename(plan_path).replace('RTPLAN_', '').replace('.dcm', '')

    # Get total MU from plan
    ds = pydicom.dcmread(plan_path)
    total_mu = 0
    for fb in ds.FractionGroupSequence:
        for rb in fb.ReferencedBeamSequence:
            if hasattr(rb, 'BeamMeterset'):
                total_mu += float(rb.BeamMeterset)

    # Find UCoMx reference by MU
    ref = None
    for r in ucomx_rows:
        if abs(r['MUs'] - total_mu) < 1:
            ref = r
            break
    if ref is None:
        print(f"{name:<16} NO MATCH (MU={total_mu:.1f})")
        continue

    ref_lsv = ref.get('LSV')
    ref_aav = ref.get('AAV')
    ref_mcs = ref.get('MCSv')
    if ref_lsv is None or ref_aav is None:
        continue

    try:
        comp = compute_metrics(plan_path)
        if comp is None:
            continue

        lsv_ratio = comp['lsv'] / ref_lsv if ref_lsv else 0
        aav_ratio = comp['aav'] / ref_aav if ref_aav else 0
        mcs_ratio = comp['mcs'] / ref_mcs if ref_mcs else 0

        print(f"{name:<16} {comp['lsv']:7.4f} {ref_lsv:7.4f} {lsv_ratio:7.4f} | "
              f"{comp['aav']:7.4f} {ref_aav:7.4f} {aav_ratio:7.4f} | "
              f"{comp['mcs']:7.4f} {ref_mcs:7.4f} {mcs_ratio:7.4f}  "
              f"min_gap={comp['min_gap']:.1f} a_max={comp['a_max']:.0f}")
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"{name:<16} ERROR: {e}")

print("\n_c = computed, _u = UCoMx reference")
