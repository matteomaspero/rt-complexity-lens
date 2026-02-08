"""Check UCoMx JA values and AAV relationship."""
import openpyxl

wb = openpyxl.load_workbook(
    r'C:\Users\teoir\OneDrive\Desktop\rt-complexity-lens\testdata\reference_dataset_v1.1\0-all-20262822356.397\dataset.xlsx'
)
ws = wb.active
headers = [c.value for c in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2):
    d = {h: row[i].value for i, h in enumerate(headers)}
    rows.append(d)

print(f"{'MUs':>12} {'JA':>10} {'AAV':>8} {'LSV':>8} {'NL':>7} {'GT':>8} {'BJAR':>8} {'PA':>8}")
for r in rows:
    bjar = r.get('BJAR', '')
    pa = r.get('PA', '')
    print(f"{r['MUs']:12.2f} {r['JA']:10.2f} {r['AAV']:8.4f} {r['LSV']:8.4f} {r['NL']:7.2f} {r['GT']:8.2f} {str(bjar):>8} {str(pa):>8}")
