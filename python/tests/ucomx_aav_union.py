"""Test AAV with per-leaf max gap (union aperture) as A_max."""
import pydicom
import numpy as np
import openpyxl
import os
import glob

# Load UCoMx reference
wb = openpyxl.load_workbook(
    r'C:\Users\teoir\OneDrive\Desktop\rt-complexity-lens\testdata\reference_dataset_v1.1\0-all-20262822356.397\dataset.xlsx'
)
ws = wb.active
headers = [c.value for c in ws[1]]
ucomx_rows = []
for row in ws.iter_rows(min_row=2):
    d = {h: row[i].value for i, h in enumerate(headers)}
    ucomx_rows.append(d)


def lsv_bank(positions, active_mask):
    idx = np.where(active_mask)[0]
    if len(idx) < 2:
        return 1.0
    diffs = np.abs(np.diff(positions[idx]))
    mx = np.max(diffs)
    return np.mean(1.0 - diffs / mx) if mx > 0 else 1.0


def compute_aav_variants(dcm_path):
    """Compute AAV with different A_max definitions."""
    ds = pydicom.dcmread(dcm_path)
    beam_results = []

    for beam in ds.BeamSequence:
        bounds = None
        n_pairs = None
        for bld in beam.BeamLimitingDeviceSequence:
            if bld.RTBeamLimitingDeviceType in ('MLCX', 'MLCY'):
                bounds = np.array(bld.LeafPositionBoundaries)
                n_pairs = bld.NumberOfLeafJawPairs
                break
        if bounds is None:
            continue

        # Parse CPs
        cp_data = []
        prev_mlc = None
        prev_y = None
        all_gaps = []
        for cp in beam.ControlPointSequence:
            mlc = None
            y = prev_y
            if hasattr(cp, 'BeamLimitingDevicePositionSequence'):
                for bldp in cp.BeamLimitingDevicePositionSequence:
                    if bldp.RTBeamLimitingDeviceType in ('MLCX', 'MLCY'):
                        pos = np.array(bldp.LeafJawPositions, dtype=float)
                        half = len(pos) // 2
                        mlc = (pos[:half], pos[half:])
                    elif bldp.RTBeamLimitingDeviceType == 'ASYMY':
                        y = np.array(bldp.LeafJawPositions, dtype=float)
            if mlc is None:
                mlc = prev_mlc
            if y is None:
                y = prev_y
            if mlc is None:
                continue

            gaps = mlc[1] - mlc[0]
            all_gaps.extend(gaps.tolist())
            w = float(cp.CumulativeMetersetWeight) if hasattr(cp, 'CumulativeMetersetWeight') else 0
            cp_data.append({
                'a': mlc[0].copy(), 'b': mlc[1].copy(),
                'gaps': gaps.copy(),
                'y': y.copy() if y is not None else np.array([-200, 200]),
                'w': w
            })
            prev_mlc = mlc
            prev_y = y

        min_gap = min(all_gaps) if all_gaps else 0
        n_ca = len(cp_data) - 1
        if n_ca == 0:
            continue

        # Build CAs with midpoint interpolation
        # Track per-leaf max gap for union aperture
        per_leaf_max_gap = np.zeros(n_pairs)
        per_leaf_max_area_contrib = np.zeros(n_pairs)  # gap × effective_width

        areas = []
        for j in range(n_ca):
            cp1, cp2 = cp_data[j], cp_data[j + 1]
            ca_a = (cp1['a'] + cp2['a']) / 2
            ca_b = (cp1['b'] + cp2['b']) / 2
            ca_gaps = ca_b - ca_a
            ca_y = (cp1['y'] + cp2['y']) / 2

            area = 0
            active = np.zeros(n_pairs, dtype=bool)
            for k in range(n_pairs):
                within_jaw = (bounds[k + 1] > ca_y[0]) and (bounds[k] < ca_y[1])
                if within_jaw and ca_gaps[k] > min_gap:
                    active[k] = True
                    eff_w = max(0, min(bounds[k + 1], ca_y[1]) - max(bounds[k], ca_y[0]))
                    contrib = ca_gaps[k] * eff_w
                    area += contrib
                    # Track per-leaf maximums
                    if ca_gaps[k] > per_leaf_max_gap[k]:
                        per_leaf_max_gap[k] = ca_gaps[k]
                    if contrib > per_leaf_max_area_contrib[k]:
                        per_leaf_max_area_contrib[k] = contrib
            areas.append(area)

        areas = np.array(areas)
        a_max_global = np.max(areas)  # max single-CP area
        a_max_union = np.sum(per_leaf_max_area_contrib)  # per-leaf max gap × width
        a_max_union_gap = 0
        # Compute union using per-leaf max gap and a fixed jaw for simplicity
        ca_y = (cp_data[0]['y'] + cp_data[-1]['y']) / 2
        for k in range(n_pairs):
            within_jaw = (bounds[k + 1] > ca_y[0]) and (bounds[k] < ca_y[1])
            if within_jaw and per_leaf_max_gap[k] > min_gap:
                eff_w = max(0, min(bounds[k + 1], ca_y[1]) - max(bounds[k], ca_y[0]))
                a_max_union_gap += per_leaf_max_gap[k] * eff_w

        mean_area = np.mean(areas)
        beam_results.append({
            'aav_global': mean_area / a_max_global if a_max_global > 0 else 0,
            'aav_union': mean_area / a_max_union if a_max_union > 0 else 0,
            'aav_union_gap': mean_area / a_max_union_gap if a_max_union_gap > 0 else 0,
            'a_max_global': a_max_global,
            'a_max_union': a_max_union,
            'a_max_union_gap': a_max_union_gap,
            'mean_area': mean_area,
        })

    if beam_results:
        n = len(beam_results)
        return {
            'aav_global': sum(r['aav_global'] for r in beam_results) / n,
            'aav_union': sum(r['aav_union'] for r in beam_results) / n,
            'aav_union_gap': sum(r['aav_union_gap'] for r in beam_results) / n,
            'a_max_global': max(r['a_max_global'] for r in beam_results),
            'a_max_union': max(r['a_max_union'] for r in beam_results),
            'mean_area': sum(r['mean_area'] for r in beam_results) / n,
        }
    return None


base = r'C:\Users\teoir\OneDrive\Desktop\rt-complexity-lens\testdata\reference_dataset_v1.1\Linac'
plans = sorted(glob.glob(os.path.join(base, '*', 'RTPLAN_*.dcm')))

print(f"{'Plan':<16} {'AAV_glob':>8} {'AAV_uni':>8} {'AAV_ref':>8} | {'ratio_g':>8} {'ratio_u':>8} | {'Amax_g':>8} {'Amax_u':>8} {'m_area':>8}")
print("-" * 110)

for plan_path in plans:
    name = os.path.basename(plan_path).replace('RTPLAN_', '').replace('.dcm', '')

    ds = pydicom.dcmread(plan_path)
    total_mu = 0
    for fb in ds.FractionGroupSequence:
        for rb in fb.ReferencedBeamSequence:
            if hasattr(rb, 'BeamMeterset'):
                total_mu += float(rb.BeamMeterset)

    ref = None
    for r in ucomx_rows:
        if abs(r['MUs'] - total_mu) < 1:
            ref = r
            break
    if ref is None:
        continue

    ref_aav = ref.get('AAV')
    if ref_aav is None:
        continue

    try:
        comp = compute_aav_variants(plan_path)
        if comp is None:
            continue

        r_g = comp['aav_global'] / ref_aav if ref_aav else 0
        r_u = comp['aav_union'] / ref_aav if ref_aav else 0

        print(f"{name:<16} {comp['aav_global']:8.4f} {comp['aav_union']:8.4f} {ref_aav:8.4f} | "
              f"{r_g:8.4f} {r_u:8.4f} | "
              f"{comp['a_max_global']:8.0f} {comp['a_max_union']:8.0f} {comp['mean_area']:8.0f}")
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"{name:<16} ERROR: {e}")
