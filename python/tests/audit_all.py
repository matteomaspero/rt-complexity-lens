#!/usr/bin/env python3
"""
End-to-end metric audit orchestrator.

Runs the three independent validation layers for every plan in
public/test-data/ and writes machine-readable artifacts that drive the
in-app Validation Report:

  1. TS ↔ Python parity          → audit_ts_python_per_plan.json
  2. ApertureComplexity (Younge) → audit_pycomplexity_per_plan.json
     (complementary metric — NOT an equivalence check)
  3. UCoMx v1.1 (MATLAB)         → audit_ucomx_per_plan.json
     (only if testdata/reference_dataset_v1.1/.../dataset.xlsx is present)

Also writes:
  - audit_summary.json
  - AUDIT_REPORT.md   (human-readable)

Usage:
    cd python && python tests/audit_all.py
"""

import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from rtplan_complexity.parser import parse_rtplan
from rtplan_complexity.metrics import calculate_plan_metrics

# Reuse identical tolerances + helper from cross_validate.py
from tests.cross_validate import (  # type: ignore
    METRIC_TOLERANCES,
    compute_python_metrics,
)

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
TEST_DATA = PROJECT_ROOT / "public" / "test-data"
REF_DIR = SCRIPT_DIR / "reference_data"
TS_REF = REF_DIR / "reference_metrics_ts.json"
PYC_REF = REF_DIR / "benchmark_pycomplexity.json"
UCOMX_XLSX = (
    PROJECT_ROOT
    / "testdata"
    / "reference_dataset_v1.1"
    / "0-all-20262822356.397"
    / "dataset.xlsx"
)


# ---------------------------------------------------------------------------
# Layer 1 — TS ↔ Python parity
# ---------------------------------------------------------------------------
def layer_ts_python(ts_plans: dict) -> dict:
    per_plan: dict = {}
    for fname, ts_metrics in ts_plans.items():
        dcm = TEST_DATA / fname
        if not dcm.exists():
            per_plan[fname] = {"status": "skipped", "reason": "file not found"}
            continue
        try:
            py = compute_python_metrics(str(dcm))
        except Exception as e:
            per_plan[fname] = {"status": "error", "reason": str(e)}
            continue
        deltas = {}
        for key, tol in METRIC_TOLERANCES.items():
            t, p = ts_metrics.get(key), py.get(key)
            if t is None or p is None:
                continue
            try:
                d = abs(float(t) - float(p))
            except (TypeError, ValueError):
                continue
            deltas[key] = {
                "ts": float(t),
                "py": float(p),
                "abs_delta": d,
                "tolerance": tol,
                "within_tol": d <= tol,
            }
        per_plan[fname] = {"status": "ok", "metrics": deltas}
    return per_plan


# ---------------------------------------------------------------------------
# Layer 2 — ApertureComplexity (Younge edge metric)
# ---------------------------------------------------------------------------
def layer_pycomplexity(ts_plans: dict) -> dict:
    if not PYC_REF.exists():
        return {"status": "skipped", "reason": f"missing {PYC_REF.name}"}
    pyc = json.loads(PYC_REF.read_text())
    per_plan = {}
    for fname, ts_metrics in ts_plans.items():
        ts_em = ts_metrics.get("EM")
        pyc_entry = pyc.get("plans", {}).get(fname)
        if ts_em is None or pyc_entry is None or "plan_edge_metric_mm_inv" not in pyc_entry:
            per_plan[fname] = {"status": "missing"}
            continue
        ci = float(pyc_entry["plan_edge_metric_mm_inv"])
        per_plan[fname] = {
            "status": "ok",
            "ours_EM_mm_inv": float(ts_em),
            "pycomplexity_CI_mm_inv": ci,
            "abs_delta": abs(float(ts_em) - ci),
            "note": (
                "Different definitions: ours EM = perimeter/(2·area); "
                "PyComplexity CI = signed Younge edge metric. "
                "Reported for transparency, NOT equivalence."
            ),
        }
    return {"status": "ok", "tool": pyc.get("tool"), "tool_url": pyc.get("tool_url"), "per_plan": per_plan}


# ---------------------------------------------------------------------------
# Layer 3 — UCoMx v1.1
# ---------------------------------------------------------------------------
def layer_ucomx(ts_plans: dict) -> dict:
    if not UCOMX_XLSX.exists():
        return {
            "status": "skipped",
            "reason": (
                "UCoMx v1.1 MATLAB reference dataset not present in this sandbox "
                f"(expected at {UCOMX_XLSX.relative_to(PROJECT_ROOT)}). "
                "Last archived run results are preserved in src/lib/validation-data.ts."
            ),
        }
    import openpyxl  # local import — only needed when xlsx exists

    from tests.cross_validate_ucomx import COMPARABLE_METRICS, compare_metric  # type: ignore

    wb = openpyxl.load_workbook(str(UCOMX_XLSX), data_only=True)
    ws_m, ws_i = wb["metrics"], wb["info"]
    headers = [str(ws_m.cell(1, c).value) for c in range(1, ws_m.max_column + 1)]
    filenames = [str(ws_i.cell(r, 2).value) for r in range(2, ws_i.max_row + 1)]
    ucomx_plans = {}
    for i, fname in enumerate(filenames):
        row = i + 2
        ucomx_plans[fname] = {
            h: ws_m.cell(row, j + 1).value for j, h in enumerate(headers)
        }

    per_plan = {}
    for fname, ts_metrics in ts_plans.items():
        u = ucomx_plans.get(fname)
        if u is None:
            per_plan[fname] = {"status": "not_in_ucomx"}
            continue
        metrics = {}
        for u_key, t_key in COMPARABLE_METRICS.items():
            uv, tv = u.get(u_key), ts_metrics.get(t_key)
            if uv is None or tv is None:
                continue
            match, delta, rel = compare_metric(uv, tv, t_key)
            if match is None:
                continue
            metrics[t_key] = {
                "ucomx_key": u_key,
                "ucomx": float(uv),
                "ts": float(tv),
                "abs_delta": float(delta),
                "rel_delta": float(rel),
                "within_tol": bool(match),
            }
        per_plan[fname] = {"status": "ok", "metrics": metrics}
    return {"status": "ok", "per_plan": per_plan}


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------
def summarize_ts_py(layer1: dict) -> dict:
    per_metric: dict = {}
    for fname, entry in layer1.items():
        if entry.get("status") != "ok":
            continue
        for k, d in entry["metrics"].items():
            per_metric.setdefault(k, []).append(d["abs_delta"])
    summary = {}
    for k, vals in per_metric.items():
        summary[k] = {
            "n": len(vals),
            "mean_abs_delta": sum(vals) / len(vals),
            "max_abs_delta": max(vals),
            "tolerance": METRIC_TOLERANCES.get(k),
            "all_within_tol": all(v <= METRIC_TOLERANCES.get(k, float("inf")) for v in vals),
        }
    return summary


def summarize_pyc(layer2: dict) -> dict:
    if layer2.get("status") != "ok":
        return {"status": layer2.get("status")}
    deltas = [
        e["abs_delta"]
        for e in layer2["per_plan"].values()
        if e.get("status") == "ok"
    ]
    return {
        "status": "ok",
        "n": len(deltas),
        "mean_abs_delta_EM_vs_CI": sum(deltas) / len(deltas) if deltas else None,
        "max_abs_delta_EM_vs_CI": max(deltas) if deltas else None,
        "note": "Different definitions; not an equivalence test.",
    }


def summarize_ucomx(layer3: dict) -> dict:
    if layer3.get("status") != "ok":
        return {"status": layer3.get("status"), "reason": layer3.get("reason")}
    per_metric: dict = {}
    outliers = []
    for fname, entry in layer3["per_plan"].items():
        if entry.get("status") != "ok":
            continue
        for k, d in entry["metrics"].items():
            per_metric.setdefault(k, []).append(d)
            if not d["within_tol"]:
                outliers.append({"plan": fname, "metric": k, **d})
    summary = {}
    for k, items in per_metric.items():
        rels = [it["rel_delta"] for it in items]
        absd = [it["abs_delta"] for it in items]
        summary[k] = {
            "n": len(items),
            "mean_abs_delta": sum(absd) / len(absd),
            "max_abs_delta": max(absd),
            "max_rel_delta_pct": max(rels) * 100,
            "all_within_tol": all(it["within_tol"] for it in items),
        }
    return {"status": "ok", "per_metric": summary, "outliers": outliers}


# ---------------------------------------------------------------------------
# Markdown report
# ---------------------------------------------------------------------------
def write_markdown(summary: dict, out_path: Path):
    ts = summary["generatedAt"]
    lines = [
        "# Audit Report",
        "",
        f"_Generated {ts}_",
        "",
        f"- Plans analyzed: **{summary['planCount']}**",
        "",
        "## Layer 1 — TS ↔ Python parity",
        "",
        "| Metric | n | Mean Δ | Max Δ | Tol | Pass |",
        "|---|---:|---:|---:|---:|:--:|",
    ]
    for k, s in sorted(summary["tsPython"].items()):
        lines.append(
            f"| {k} | {s['n']} | {s['mean_abs_delta']:.6f} | "
            f"{s['max_abs_delta']:.6f} | {s['tolerance']} | "
            f"{'✓' if s['all_within_tol'] else '✗'} |"
        )

    lines += ["", "## Layer 2 — ApertureComplexity (Younge edge metric)"]
    pyc = summary["pycomplexity"]
    if pyc.get("status") == "ok":
        lines += [
            "",
            f"- Plans compared: {pyc['n']}",
            f"- Mean |Δ(EM, CI)|: {pyc['mean_abs_delta_EM_vs_CI']:.4f} mm⁻¹",
            f"- Max  |Δ(EM, CI)|: {pyc['max_abs_delta_EM_vs_CI']:.4f} mm⁻¹",
            "",
            "> Different formulations — reported for transparency, NOT equivalence.",
        ]
    else:
        lines.append(f"- Status: **skipped** ({pyc.get('status')})")

    lines += ["", "## Layer 3 — UCoMx v1.1 (MATLAB)"]
    uc = summary["ucomx"]
    if uc.get("status") == "ok":
        lines += [
            "",
            "| Metric | n | Mean abs Δ | Max abs Δ | Max rel % | Pass |",
            "|---|---:|---:|---:|---:|:--:|",
        ]
        for k, s in sorted(uc["per_metric"].items()):
            lines.append(
                f"| {k} | {s['n']} | {s['mean_abs_delta']:.6f} | "
                f"{s['max_abs_delta']:.6f} | {s['max_rel_delta_pct']:.2f}% | "
                f"{'✓' if s['all_within_tol'] else '✗'} |"
            )
        if uc.get("outliers"):
            lines += ["", "### Outliers"]
            for o in uc["outliers"]:
                lines.append(
                    f"- `{o['plan']}` `{o['metric']}`: UCoMx={o['ucomx']:.4f} "
                    f"TS={o['ts']:.4f} (Δ={o['abs_delta']:.4f}, "
                    f"{o['rel_delta']*100:.2f}%)"
                )
    else:
        lines += [
            "",
            f"- Status: **skipped** — {uc.get('reason')}",
        ]

    out_path.write_text("\n".join(lines) + "\n")


# ---------------------------------------------------------------------------
def main():
    if not TS_REF.exists():
        print(f"ERROR: missing {TS_REF}. Regenerate with the TS exporter test.")
        sys.exit(1)

    ts_data = json.loads(TS_REF.read_text())
    ts_plans = ts_data["plans"]
    print(f"Auditing {len(ts_plans)} plans …")

    layer1 = layer_ts_python(ts_plans)
    print("  Layer 1 (TS↔Python)  done")
    layer2 = layer_pycomplexity(ts_plans)
    print(f"  Layer 2 (ApertureComplexity)  {layer2.get('status', 'ok')}")
    layer3 = layer_ucomx(ts_plans)
    print(f"  Layer 3 (UCoMx)  {layer3.get('status', 'ok')}")

    REF_DIR.mkdir(parents=True, exist_ok=True)
    (REF_DIR / "audit_ts_python_per_plan.json").write_text(json.dumps(layer1, indent=2))
    (REF_DIR / "audit_pycomplexity_per_plan.json").write_text(json.dumps(layer2, indent=2))
    (REF_DIR / "audit_ucomx_per_plan.json").write_text(json.dumps(layer3, indent=2))

    summary = {
        "generatedAt": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "planCount": len(ts_plans),
        "tsPython": summarize_ts_py(layer1),
        "pycomplexity": summarize_pyc(layer2),
        "ucomx": summarize_ucomx(layer3),
    }
    (REF_DIR / "audit_summary.json").write_text(json.dumps(summary, indent=2))
    write_markdown(summary, REF_DIR / "AUDIT_REPORT.md")
    print(f"\nWrote artifacts to {REF_DIR.relative_to(PROJECT_ROOT)}/")
    print("  - audit_ts_python_per_plan.json")
    print("  - audit_pycomplexity_per_plan.json")
    print("  - audit_ucomx_per_plan.json")
    print("  - audit_summary.json")
    print("  - AUDIT_REPORT.md")


if __name__ == "__main__":
    main()
