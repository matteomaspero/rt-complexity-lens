#!/usr/bin/env python3
"""
Cross-validate TypeScript metrics against UCoMx v1.1 MATLAB reference output.

Reads:
  - UCoMx xlsx output from testdata/reference_dataset_v1.1/0-all-*/dataset.xlsx
  - TS reference JSON from python/tests/reference_data/reference_metrics_ts.json

Reports metric-by-metric comparison for all overlapping plans.
"""

import json
import sys
from pathlib import Path

import openpyxl

# ============================================================================
# Paths
# ============================================================================
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
UCOMX_DIR = PROJECT_ROOT / "testdata" / "reference_dataset_v1.1" / "0-all-20262822356.397"
UCOMX_XLSX = UCOMX_DIR / "dataset.xlsx"
TS_REF_JSON = SCRIPT_DIR / "reference_data" / "reference_metrics_ts.json"

# ============================================================================
# UCoMx → TS metric name mapping
# ============================================================================
# UCoMx name → TS name  (None = no TS equivalent yet)
UCOMX_TO_TS = {
    # Deliverability / core
    "MUs":              "totalMU",
    "MUCA":             "MUCA",
    "MUcGy":            None,           # MU per cGy – we don't compute this
    "MD":               "MD",
    "PMU":              None,           # Plan MU (?)  – check meaning
    "LT":               "LT",
    "LTMU":             "LTMU",
    "LTNLMU":           "LTNLMU",
    "LTNL":             None,           # Leaf Travel per Leaf (no MU norm)
    "LNA":              "LNA",
    "MUdeg":            "MUperDegree",
    "LTAL":             "LTAL",
    "mDRV":             "mDRV",
    "mGSV":             "mGSV",
    "DR":               "avgDoseRate",
    "GS":               "GS",
    "LS":               "LS",
    "MCSv":             "MCS",          # MCSv = MCS (v for "volume"/plan-level?)
    "AAV":              "AAV",
    "LSV":              "LSV",
    "TG":               "TG",
    "MIt":              None,           # MI total – we compute MI differently?
    "MIs":              None,           # MI segment
    "MIa":              None,           # MI aperture
    "PI":               "PI",
    # Info / basics
    "PrescribedDose":   None,           # Not a complexity metric per se
    "NL":               None,           # Number of leaves
    "AL":               None,           # Arc Length per arc?
    "CAL":              None,           # Control point Arc Length?
    "GT":               "GT",
    "NArcs":            None,           # Number of arcs (we have beamCount)
    "NSeg":             None,           # Number of segments (we have numberOfControlPoints)
    "dt":               None,           # Delivery time – may map to totalDeliveryTime
    # Accuracy
    "PA":               "PA",
    "ALG":              "LG",           # Average Leaf Gap = our LG
    "psmall_10mm":      "psmall",       # Our psmall uses which threshold?
    "psmall_20mm":      None,
    "psmall_30mm":      None,
    "EFS":              "EFS",
    "SAS5mm":           "SAS5",
    "SAS10mm":          "SAS10",
    "SAS25mm":          None,           # We don't compute SAS25
    "SAS50mm":          None,           # We don't compute SAS50
    "EM":               "EM",
    "BJAR":             None,           # BEV Jaw Area Ratio – not implemented
    "MAD":              "MAD",
    "P":                "PM",           # Plan Modulation = 1 - MCS ?
    "JA":               "JA",
}

# List of metrics we expect to exactly match (same algorithm)
COMPARABLE_METRICS = {k: v for k, v in UCOMX_TO_TS.items() if v is not None}

# ============================================================================
# Tolerance thresholds per metric (relative or absolute)
# ============================================================================
# Some metrics may have slight differences due to:
#   - Floating point precision
#   - MLC bank ordering interpretation
#   - Rounding in MU-weighting
#   - Different handling of first/last control point
RELATIVE_TOL = 0.05   # 5% relative tolerance
ABSOLUTE_TOL = {       # metric-specific absolute tolerances
    "LG": 0.5,        # mm – rounding sensitivity
    "MAD": 0.5,       # mm
    "EFS": 1.0,       # mm
    "EM": 0.01,
    "mDRV": 1.0,      # MU/min – delivery estimation differences
    "mGSV": 0.05,     # deg/s
    "avgDoseRate": 10, # MU/min
}


def load_ucomx_data():
    """Load UCoMx xlsx and return {filename: {metric: value}}."""
    wb = openpyxl.load_workbook(str(UCOMX_XLSX), data_only=True)
    ws_m = wb["metrics"]
    ws_i = wb["info"]

    # Parse headers from metrics sheet
    headers = []
    for c in range(1, ws_m.max_column + 1):
        headers.append(str(ws_m.cell(1, c).value))

    # Parse filenames from info sheet
    filenames = []
    for r in range(2, ws_i.max_row + 1):
        filenames.append(str(ws_i.cell(r, 2).value))

    # Build plan dict
    plans = {}
    for i, fname in enumerate(filenames):
        row_idx = i + 2  # 1-indexed, skip header
        metrics = {}
        for j, header in enumerate(headers):
            val = ws_m.cell(row_idx, j + 1).value
            if val is not None:
                metrics[header] = val
        plans[fname] = metrics

    return plans


def load_ts_data():
    """Load TS reference JSON and return {filename: {metric: value}}."""
    with open(str(TS_REF_JSON)) as f:
        data = json.load(f)
    return data["plans"]


def compare_metric(ucomx_val, ts_val, ts_key):
    """Compare two metric values. Return (match, delta, rel_delta)."""
    if ucomx_val is None or ts_val is None:
        return None, None, None

    try:
        u = float(ucomx_val)
        t = float(ts_val)
    except (ValueError, TypeError):
        return None, None, None

    # Handle zeros
    if u == 0 and t == 0:
        return True, 0.0, 0.0

    delta = abs(u - t)

    # Relative delta (relative to UCoMx reference)
    denom = max(abs(u), abs(t), 1e-10)
    rel_delta = delta / denom

    # Check absolute tolerance
    abs_tol = ABSOLUTE_TOL.get(ts_key, 0.01)
    if delta <= abs_tol:
        return True, delta, rel_delta

    # Check relative tolerance
    if rel_delta <= RELATIVE_TOL:
        return True, delta, rel_delta

    return False, delta, rel_delta


def main():
    print("=" * 80)
    print("UCoMx v1.1 (MATLAB) vs TypeScript Cross-Validation")
    print("=" * 80)

    ucomx_plans = load_ucomx_data()
    ts_plans = load_ts_data()

    print(f"\nUCoMx plans: {len(ucomx_plans)}")
    print(f"TS plans:    {len(ts_plans)}")

    # Find overlapping plans
    overlap = set(ucomx_plans.keys()) & set(ts_plans.keys())
    print(f"Overlapping: {len(overlap)}")
    print()

    if not overlap:
        print("No overlapping plans found!")
        # Show available names for debugging
        print("\nUCoMx filenames:", sorted(ucomx_plans.keys()))
        print("\nTS filenames:", sorted(ts_plans.keys()))
        sys.exit(1)

    # Track results
    all_results = []  # (plan, ucomx_key, ts_key, ucomx_val, ts_val, match, delta, rel_delta)
    summary_by_metric = {}  # ts_key -> {match, mismatch, missing}

    for plan in sorted(overlap):
        u_metrics = ucomx_plans[plan]
        t_metrics = ts_plans[plan]

        for ucomx_key, ts_key in COMPARABLE_METRICS.items():
            if ts_key not in summary_by_metric:
                summary_by_metric[ts_key] = {"match": 0, "mismatch": 0, "missing": 0, "deltas": []}

            ucomx_val = u_metrics.get(ucomx_key)
            ts_val = t_metrics.get(ts_key)

            if ucomx_val is None or ts_val is None:
                summary_by_metric[ts_key]["missing"] += 1
                all_results.append((plan, ucomx_key, ts_key, ucomx_val, ts_val, None, None, None))
                continue

            match, delta, rel_delta = compare_metric(ucomx_val, ts_val, ts_key)
            if match is None:
                summary_by_metric[ts_key]["missing"] += 1
            elif match:
                summary_by_metric[ts_key]["match"] += 1
                summary_by_metric[ts_key]["deltas"].append(rel_delta)
            else:
                summary_by_metric[ts_key]["mismatch"] += 1
                summary_by_metric[ts_key]["deltas"].append(rel_delta)

            all_results.append((plan, ucomx_key, ts_key, ucomx_val, ts_val, match, delta, rel_delta))

    # ========================================================================
    # Print per-metric summary
    # ========================================================================
    print("=" * 80)
    print("PER-METRIC SUMMARY (across all overlapping plans)")
    print("=" * 80)
    print(f"{'TS Key':<16} {'UCoMx Key':<14} {'Match':>6} {'Mismatch':>9} {'Missing':>8} {'Max Rel D':>10}")
    print("-" * 80)

    mismatch_details = []
    for ts_key in sorted(summary_by_metric.keys()):
        s = summary_by_metric[ts_key]
        # Find corresponding UCoMx key
        ucomx_key = [k for k, v in COMPARABLE_METRICS.items() if v == ts_key][0]
        max_rel = max(s["deltas"]) * 100 if s["deltas"] else 0
        status = "PASS" if s["mismatch"] == 0 else "FAIL"
        print(f"{ts_key:<16} {ucomx_key:<14} {s['match']:>6} {s['mismatch']:>9} {s['missing']:>8} {max_rel:>9.2f}%  {status}")

        if s["mismatch"] > 0:
            mismatch_details.append(ts_key)

    # ========================================================================
    # Print mismatches in detail
    # ========================================================================
    if mismatch_details:
        print("\n" + "=" * 80)
        print("MISMATCH DETAILS")
        print("=" * 80)
        for plan, ucomx_key, ts_key, u_val, t_val, match, delta, rel_delta in all_results:
            if match is False:
                print(f"  {plan:<40} {ts_key:<12} UCoMx={u_val:<14.6f} TS={t_val:<14.6f} D={delta:<10.4f} rel={rel_delta*100:.2f}%")

    # ========================================================================
    # Print metrics NOT in TS
    # ========================================================================
    print("\n" + "=" * 80)
    print("UCoMx METRICS NOT MAPPED TO TS")
    print("=" * 80)
    for ucomx_key, ts_key in sorted(UCOMX_TO_TS.items()):
        if ts_key is None:
            # Get sample value
            sample_plan = list(overlap)[0]
            sample_val = ucomx_plans[sample_plan].get(ucomx_key)
            print(f"  {ucomx_key:<16} sample={sample_val}")

    # ========================================================================
    # Print TS metrics not in UCoMx
    # ========================================================================
    print("\n" + "=" * 80)
    print("TS METRICS NOT IN UCoMx")
    print("=" * 80)
    ts_mapped = set(UCOMX_TO_TS.values()) - {None}
    sample_plan = list(overlap)[0]
    for ts_key in sorted(ts_plans[sample_plan].keys()):
        if ts_key not in ts_mapped:
            print(f"  {ts_key:<20} = {ts_plans[sample_plan].get(ts_key)}")

    # ========================================================================
    # Per-plan detailed comparison for one sample plan
    # ========================================================================
    print("\n" + "=" * 80)
    sample = sorted(overlap)[0]
    print(f"DETAILED COMPARISON: {sample}")
    print("=" * 80)
    u_m = ucomx_plans[sample]
    t_m = ts_plans[sample]
    print(f"{'Metric':<16} {'UCoMx Key':<14} {'UCoMx Value':>14} {'TS Value':>14} {'Delta':>10} {'Rel%':>8} {'Status':>8}")
    print("-" * 90)
    for ucomx_key, ts_key in sorted(COMPARABLE_METRICS.items(), key=lambda x: x[1]):
        u_val = u_m.get(ucomx_key)
        t_val = t_m.get(ts_key)
        if u_val is not None and t_val is not None:
            match, delta, rel_delta = compare_metric(u_val, t_val, ts_key)
            status = "PASS" if match else "FAIL"
            print(f"{ts_key:<16} {ucomx_key:<14} {float(u_val):>14.6f} {float(t_val):>14.6f} {delta:>10.4f} {rel_delta*100:>7.2f}% {status:>8}")
        elif u_val is not None:
            print(f"{ts_key:<16} {ucomx_key:<14} {float(u_val):>14.6f} {'N/A':>14} {'':>10} {'':>8} {'MISSING':>8}")
        elif t_val is not None:
            print(f"{ts_key:<16} {ucomx_key:<14} {'N/A':>14} {float(t_val):>14.6f} {'':>10} {'':>8} {'EXTRA':>8}")
        else:
            print(f"{ts_key:<16} {ucomx_key:<14} {'N/A':>14} {'N/A':>14} {'':>10} {'':>8} {'BOTH N/A':>8}")

    # Final summary
    total_matches = sum(s["match"] for s in summary_by_metric.values())
    total_mismatches = sum(s["mismatch"] for s in summary_by_metric.values())
    total_missing = sum(s["missing"] for s in summary_by_metric.values())
    print(f"\n{'='*80}")
    print(f"TOTAL: {total_matches} matches, {total_mismatches} mismatches, {total_missing} missing")
    print(f"{'='*80}")

    return 0 if total_mismatches == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
